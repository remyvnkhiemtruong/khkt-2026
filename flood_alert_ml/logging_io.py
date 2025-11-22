from __future__ import annotations

import csv
from typing import Dict, Any

from openpyxl import Workbook, load_workbook

from .config import CSV_LOG_PATH, XLSX_LOG_PATH, LOG_COLUMNS
from .utils import atomic_write


class CSVLogger:
    def append(self, row: Dict[str, Any]) -> None:
        first = not CSV_LOG_PATH.exists()
        with atomic_write(CSV_LOG_PATH) as tmp:
            # copy old file if exists
            if CSV_LOG_PATH.exists():
                tmp.write_text(CSV_LOG_PATH.read_text(encoding="utf-8"), encoding="utf-8")
            with open(tmp, "a", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=LOG_COLUMNS)
                if first:
                    w.writeheader()
                w.writerow({k: row.get(k, "") for k in LOG_COLUMNS})


class ExcelLogger:
    def append(self, row: Dict[str, Any]) -> None:
        if XLSX_LOG_PATH.exists():
            wb = load_workbook(XLSX_LOG_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(LOG_COLUMNS)
        ws.append([row.get(k, "") for k in LOG_COLUMNS])
        wb.save(XLSX_LOG_PATH)
